"""Seed the database with equipment data from the Master sheet of the Excel file."""
import openpyxl
from models import db, Equipment
from config import Config


def seed_from_excel(app):
    """Read the Master sheet and populate equipment table."""
    with app.app_context():
        # Check if already seeded
        if Equipment.query.count() > 0:
            print("Database already seeded. Skipping.")
            return

        wb = openpyxl.load_workbook(Config.EXCEL_PATH, data_only=True)

        # --- Seed Equipment from Master sheet ---
        ws = wb['Master']
        current_section = None
        current_name = None

        seen_ids = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=7, values_only=True):
            section, name, line, eq_id, criticality, status, overdue = row

            if section:
                current_section = section
            if name:
                current_name = name

            if eq_id and eq_id not in seen_ids:
                seen_ids.add(eq_id)
                equipment = Equipment(
                    equipment_id=eq_id,
                    factory_section=current_section,
                    equipment_name=current_name,
                    line=line,
                    equipment_criticality=criticality,
                    overall_pm_status=status or 'Green',
                    overdue_pm_count=overdue or 0
                )
                db.session.add(equipment)

        db.session.commit()
        count = Equipment.query.count()
        print(f"Seeded {count} equipment records from Master sheet.")
        wb.close()
