"""Flask API routes for PM Tracker."""
from flask import Blueprint, request, jsonify
from models import db, Equipment, PMTask, PMHistory
from datetime import datetime, date, timedelta
from sqlalchemy import or_
import openpyxl
import os
import tempfile

api = Blueprint('api', __name__)


# ==================== AUTH ROUTES ====================

# Simple admin credentials (in production, use proper auth)
ADMIN_CREDENTIALS = {'username': 'admin', 'password': 'admin123'}

@api.route('/login', methods=['POST'])
def login():
    """Simple login - returns role."""
    data = request.get_json()
    username = data.get('username', '')
    password = data.get('password', '')
    if username == ADMIN_CREDENTIALS['username'] and password == ADMIN_CREDENTIALS['password']:
        return jsonify({'role': 'admin', 'username': username, 'message': 'Login successful'})
    # Any other credentials get viewer role
    if username.strip():
        return jsonify({'role': 'viewer', 'username': username, 'message': 'Login successful (view-only)'})
    return jsonify({'error': 'Please enter a username'}), 401


# ==================== DROPDOWN OPTIONS ====================

@api.route('/dropdown-options', methods=['GET'])
def get_dropdown_options():
    """Return predefined dropdown options for forms."""
    return jsonify({
        'spare_parts': [
            'Belt', 'Bearing', 'Motor', 'Sensor', 'Filter', 'Gasket', 'Seal',
            'Roller', 'Conveyor Belt', 'Heating Element', 'Thermocouple',
            'Pressure Gauge', 'Valve', 'Pump', 'Fuse', 'Relay', 'Contactor',
            'Spring', 'Gear', 'Chain', 'Nozzle', 'Blade', 'Brush',
            'Pneumatic Cylinder', 'Solenoid Valve', 'Drive Belt',
            'Glass Sheet', 'Suction Cup', 'Gripper Pad', 'PCB Board',
            'Ribbon', 'Flux Pen', 'Soldering Tip', 'IR Lamp', 'UV Lamp',
            'O-Ring', 'Coupling', 'Shaft', 'Pulley', 'Encoder'
        ],
        'consumables': [
            'Lubricant', 'Grease', 'Cleaning Solvent', 'Wipes', 'Cloth',
            'Adhesive Tape', 'Isopropyl Alcohol', 'Compressed Air',
            'Silicon Spray', 'Thread Lock', 'Anti-Seize Compound',
            'Contact Cleaner', 'Degreaser', 'Flux', 'Solder Wire',
            'Coolant', 'Hydraulic Oil', 'Machine Oil', 'Cutting Fluid',
            'Masking Tape', 'Cable Ties', 'Gloves', 'Safety Glasses',
            'Sandpaper', 'Cotton Swab', 'Thermal Paste'
        ],
        'done_by': [
            'Rajesh Kumar', 'Amit Singh', 'Suresh Patel', 'Vikram Sharma',
            'Mohammed Ali', 'Ravi Verma', 'Sanjay Gupta', 'Deepak Yadav',
            'Manoj Tiwari', 'Arun Mishra', 'Pradeep Joshi', 'Ramesh Rao'
        ],
        'verified_by': [
            'Supervisor - Rajiv', 'Supervisor - Anand', 'Supervisor - Prakash',
            'Manager - Sunil', 'Manager - Vinod', 'QC - Ashok', 'QC - Suman',
            'Shift Lead - Mohan', 'Shift Lead - Ganesh'
        ]
    })


# ==================== PREDICTION ROUTES ====================

@api.route('/predictions', methods=['GET'])
def get_predictions():
    """Get PM task predictions - upcoming due, overdue forecast, workload."""
    today = date.today()
    tasks = PMTask.query.filter(PMTask.next_due_date.isnot(None)).all()

    upcoming_7 = []   # Due in next 7 days
    upcoming_30 = []   # Due in next 30 days
    overdue_now = []   # Currently overdue
    workload = {}      # Daily workload forecast

    for task in tasks:
        days_diff = (task.next_due_date - today).days
        eq = Equipment.query.get(task.equipment_db_id)
        eq_id = eq.equipment_id if eq else 'Unknown'
        eq_name = eq.equipment_name if eq else 'Unknown'
        section = eq.factory_section if eq else ''
        criticality = eq.equipment_criticality if eq else ''

        task_info = {
            'task_id': task.id,
            'equipment_id': eq_id,
            'equipment_name': eq_name,
            'equipment_db_id': task.equipment_db_id,
            'section': section,
            'criticality': criticality,
            'task_no': task.task_no,
            'description': task.pm_task_description,
            'frequency_days': task.frequency_days,
            'next_due_date': task.next_due_date.isoformat(),
            'days_until_due': days_diff,
            'status': task.status or task.compute_status()
        }

        if days_diff < 0:
            overdue_now.append(task_info)
        elif days_diff <= 7:
            upcoming_7.append(task_info)
        elif days_diff <= 30:
            upcoming_30.append(task_info)

        # Build daily workload for next 30 days
        if 0 <= days_diff <= 30:
            day_key = task.next_due_date.isoformat()
            workload[day_key] = workload.get(day_key, 0) + 1

    # Sort
    overdue_now.sort(key=lambda x: x['days_until_due'])
    upcoming_7.sort(key=lambda x: x['days_until_due'])
    upcoming_30.sort(key=lambda x: x['days_until_due'])

    # Recurring prediction - tasks that will recur in next 90 days
    recurring = []
    for task in tasks:
        if task.frequency_days and task.frequency_days > 0:
            next_dates = []
            d = task.next_due_date
            for _ in range(5):
                if d and (d - today).days <= 90:
                    next_dates.append(d.isoformat())
                if d and task.frequency_days:
                    d = d + timedelta(days=task.frequency_days)
                else:
                    break
            eq = Equipment.query.get(task.equipment_db_id)
            if next_dates:
                recurring.append({
                    'equipment_id': eq.equipment_id if eq else 'Unknown',
                    'description': task.pm_task_description,
                    'frequency_days': task.frequency_days,
                    'upcoming_dates': next_dates
                })

    return jsonify({
        'overdue_now': overdue_now,
        'upcoming_7_days': upcoming_7,
        'upcoming_30_days': upcoming_30,
        'workload_forecast': workload,
        'recurring_predictions': recurring[:20],
        'summary': {
            'total_overdue': len(overdue_now),
            'due_this_week': len(upcoming_7),
            'due_this_month': len(upcoming_30),
            'total_tasks_tracked': len(tasks)
        }
    })


# ==================== EQUIPMENT ROUTES ====================

@api.route('/equipment', methods=['GET'])
def get_equipment():
    """Get all equipment with optional filters."""
    section = request.args.get('section')
    line = request.args.get('line')
    name = request.args.get('name')
    status = request.args.get('status')
    search = request.args.get('search')

    query = Equipment.query

    if section:
        query = query.filter(Equipment.factory_section == section)
    if line:
        query = query.filter(Equipment.line == line)
    if name:
        query = query.filter(Equipment.equipment_name == name)
    if status:
        query = query.filter(Equipment.overall_pm_status == status)
    if search:
        query = query.filter(
            or_(
                Equipment.equipment_id.ilike(f'%{search}%'),
                Equipment.equipment_name.ilike(f'%{search}%')
            )
        )

    equipment_list = query.order_by(Equipment.id).all()
    return jsonify([e.to_dict() for e in equipment_list])


@api.route('/equipment/<int:id>', methods=['GET'])
def get_equipment_by_id(id):
    """Get single equipment by database ID."""
    equipment = Equipment.query.get_or_404(id)
    data = equipment.to_dict()
    data['tasks'] = [t.to_dict() for t in equipment.tasks]
    return jsonify(data)


@api.route('/equipment/by-eid/<equipment_id>', methods=['GET'])
def get_equipment_by_eid(equipment_id):
    """Get single equipment by equipment_id string."""
    equipment = Equipment.query.filter_by(equipment_id=equipment_id).first_or_404()
    data = equipment.to_dict()
    data['tasks'] = [t.to_dict() for t in equipment.tasks]
    return jsonify(data)


@api.route('/equipment', methods=['POST'])
def create_equipment():
    """Create new equipment."""
    data = request.get_json()
    equipment = Equipment(
        equipment_id=data['equipment_id'],
        factory_section=data.get('factory_section', ''),
        equipment_name=data.get('equipment_name', ''),
        line=data.get('line', ''),
        equipment_criticality=data.get('equipment_criticality', ''),
        overall_pm_status='Green',
        overdue_pm_count=0
    )
    db.session.add(equipment)
    db.session.commit()
    return jsonify(equipment.to_dict()), 201


@api.route('/equipment/<int:id>', methods=['PUT'])
def update_equipment(id):
    """Update equipment details."""
    equipment = Equipment.query.get_or_404(id)
    data = request.get_json()

    for field in ['factory_section', 'equipment_name', 'line', 'equipment_criticality']:
        if field in data:
            setattr(equipment, field, data[field])

    db.session.commit()
    return jsonify(equipment.to_dict())


@api.route('/equipment/<int:id>', methods=['DELETE'])
def delete_equipment(id):
    """Delete equipment and all its tasks."""
    equipment = Equipment.query.get_or_404(id)
    db.session.delete(equipment)
    db.session.commit()
    return jsonify({'message': f'Equipment {equipment.equipment_id} deleted'})


# ==================== PM TASK ROUTES ====================

@api.route('/equipment/<int:equipment_id>/tasks', methods=['GET'])
def get_tasks(equipment_id):
    """Get all PM tasks for an equipment."""
    equipment = Equipment.query.get_or_404(equipment_id)
    tasks = PMTask.query.filter_by(equipment_db_id=equipment_id).order_by(PMTask.task_no).all()
    return jsonify([t.to_dict() for t in tasks])


@api.route('/equipment/<int:equipment_id>/tasks', methods=['POST'])
def create_task(equipment_id):
    """Create a new PM task for an equipment."""
    equipment = Equipment.query.get_or_404(equipment_id)
    data = request.get_json()

    # Auto-assign task number
    max_task = db.session.query(db.func.max(PMTask.task_no)).filter_by(
        equipment_db_id=equipment_id
    ).scalar() or 0

    task = PMTask(
        equipment_db_id=equipment_id,
        task_no=max_task + 1,
        pm_task_description=data.get('pm_task_description', ''),
        frequency_days=data.get('frequency_days'),
        tolerance_days=data.get('tolerance_days', 0),
        next_due_date=datetime.strptime(data['next_due_date'], '%Y-%m-%d').date() if data.get('next_due_date') else None,
        status='Pending',
        consumables=data.get('consumables', ''),
        spare_parts=data.get('spare_parts', ''),
        done_by=data.get('done_by', ''),
        verified_by=data.get('verified_by', ''),
        remarks=data.get('remarks', '')
    )
    db.session.add(task)

    # Recalculate equipment status
    _recalculate_equipment_status(equipment)

    db.session.commit()
    return jsonify(task.to_dict()), 201


@api.route('/tasks/<int:task_id>', methods=['PUT'])
def update_task(task_id):
    """Update a PM task."""
    task = PMTask.query.get_or_404(task_id)
    data = request.get_json()

    if 'pm_task_description' in data:
        task.pm_task_description = data['pm_task_description']
    if 'frequency_days' in data:
        task.frequency_days = data['frequency_days']
    if 'tolerance_days' in data:
        task.tolerance_days = data['tolerance_days']
    if 'next_due_date' in data:
        task.next_due_date = datetime.strptime(data['next_due_date'], '%Y-%m-%d').date() if data['next_due_date'] else None
    if 'consumables' in data:
        task.consumables = data['consumables']
    if 'spare_parts' in data:
        task.spare_parts = data['spare_parts']
    if 'done_by' in data:
        task.done_by = data['done_by']
    if 'verified_by' in data:
        task.verified_by = data['verified_by']
    if 'remarks' in data:
        task.remarks = data['remarks']

    task.status = task.compute_status()

    # Recalculate equipment status
    equipment = Equipment.query.get(task.equipment_db_id)
    _recalculate_equipment_status(equipment)

    db.session.commit()
    return jsonify(task.to_dict())


@api.route('/tasks/<int:task_id>', methods=['DELETE'])
def delete_task(task_id):
    """Delete a PM task."""
    task = PMTask.query.get_or_404(task_id)
    equipment = Equipment.query.get(task.equipment_db_id)
    db.session.delete(task)
    _recalculate_equipment_status(equipment)
    db.session.commit()
    return jsonify({'message': 'Task deleted'})


@api.route('/tasks/<int:task_id>/complete', methods=['POST'])
def complete_task(task_id):
    """Mark a PM task as completed - moves to history and reschedules next due date."""
    task = PMTask.query.get_or_404(task_id)
    data = request.get_json() or {}
    equipment = Equipment.query.get(task.equipment_db_id)

    actual_done_date_str = data.get('actual_done_date', date.today().isoformat())
    actual_done_date = datetime.strptime(actual_done_date_str, '%Y-%m-%d').date()

    # Save to PM History
    history = PMHistory(
        equipment_id=equipment.equipment_id,
        task_no=task.task_no,
        pm_task=task.pm_task_description,
        frequency_days=task.frequency_days,
        tolerance_days=task.tolerance_days,
        old_next_due_date=task.next_due_date,
        actual_done_date=actual_done_date,
        consumables=data.get('consumables', task.consumables),
        spare_parts=data.get('spare_parts', task.spare_parts),
        done_by=data.get('done_by', task.done_by),
        verified_by=data.get('verified_by', task.verified_by),
        remarks=data.get('remarks', task.remarks),
        completion_timestamp=datetime.utcnow()
    )
    db.session.add(history)

    # Reschedule: next due date = actual done date + frequency
    if task.frequency_days:
        task.next_due_date = actual_done_date + timedelta(days=task.frequency_days)
    task.actual_done_date = actual_done_date
    task.status = 'Done'
    task.consumables = data.get('consumables', task.consumables)
    task.spare_parts = data.get('spare_parts', task.spare_parts)
    task.done_by = data.get('done_by', task.done_by)
    task.verified_by = data.get('verified_by', task.verified_by)
    task.remarks = data.get('remarks', task.remarks)

    # Reset for next cycle after saving
    task.actual_done_date = None
    task.status = 'Pending'

    _recalculate_equipment_status(equipment)
    db.session.commit()

    return jsonify({
        'task': task.to_dict(),
        'history': history.to_dict(),
        'message': f'Task completed. Next due: {task.next_due_date}'
    })


# ==================== PM HISTORY ROUTES ====================

@api.route('/history', methods=['GET'])
def get_history():
    """Get PM history with optional filters."""
    equipment_id = request.args.get('equipment_id')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)

    query = PMHistory.query

    if equipment_id:
        query = query.filter(PMHistory.equipment_id == equipment_id)

    pagination = query.order_by(PMHistory.completion_timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'items': [h.to_dict() for h in pagination.items],
        'total': pagination.total,
        'page': pagination.page,
        'pages': pagination.pages,
        'per_page': per_page
    })


# ==================== DASHBOARD ROUTES ====================

@api.route('/dashboard', methods=['GET'])
def get_dashboard():
    """Get dashboard statistics - mirrors the Dashboard sheet."""
    total_equipment = Equipment.query.count()

    # Count by PM status
    green_count = Equipment.query.filter_by(overall_pm_status='Green').count()
    orange_count = Equipment.query.filter_by(overall_pm_status='Orange').count()
    red_count = Equipment.query.filter_by(overall_pm_status='Red').count()

    # Overdue tasks breakdown
    today = date.today()
    all_tasks = PMTask.query.filter(
        PMTask.next_due_date < today,
        PMTask.status != 'Done'
    ).all()

    overdue_1_3 = 0
    overdue_4_7 = 0
    overdue_gt_7 = 0
    critical_overdue = 0

    for task in all_tasks:
        days_overdue = (today - task.next_due_date).days
        if 1 <= days_overdue <= 3:
            overdue_1_3 += 1
        elif 4 <= days_overdue <= 7:
            overdue_4_7 += 1
        elif days_overdue > 7:
            overdue_gt_7 += 1

    # Critical equipment with overdue tasks
    critical_equipment_ids = db.session.query(PMTask.equipment_db_id).filter(
        PMTask.next_due_date < today,
        PMTask.status != 'Done'
    ).distinct().all()
    critical_overdue = len(critical_equipment_ids)

    # Section-wise breakdown
    sections = db.session.query(
        Equipment.factory_section,
        db.func.count(Equipment.id)
    ).group_by(Equipment.factory_section).all()

    # Equipment name breakdown
    equipment_types = db.session.query(
        Equipment.equipment_name,
        db.func.count(Equipment.id)
    ).group_by(Equipment.equipment_name).all()

    # Line breakdown
    lines = db.session.query(
        Equipment.line,
        db.func.count(Equipment.id)
    ).group_by(Equipment.line).all()

    return jsonify({
        'total_equipment': total_equipment,
        'status_summary': {
            'green': green_count,
            'orange': orange_count,
            'red': red_count
        },
        'overdue_aging': {
            'days_1_3': overdue_1_3,
            'days_4_7': overdue_4_7,
            'days_gt_7': overdue_gt_7
        },
        'critical_equipment_overdue': critical_overdue,
        'sections': {s[0]: s[1] for s in sections if s[0]},
        'equipment_types': {e[0]: e[1] for e in equipment_types if e[0]},
        'lines': {l[0]: l[1] for l in lines if l[0]},
        'total_tasks': PMTask.query.count(),
        'total_completed': PMHistory.query.count(),
        'overdue_equipment_list': _get_overdue_equipment_list()
    })


@api.route('/dashboard/recalculate', methods=['POST'])
def recalculate_all_statuses():
    """Recalculate PM status for all equipment."""
    equipment_list = Equipment.query.all()
    for eq in equipment_list:
        _recalculate_equipment_status(eq)
    db.session.commit()
    return jsonify({'message': f'Recalculated status for {len(equipment_list)} equipment'})


# ==================== METADATA ROUTES ====================

@api.route('/sections', methods=['GET'])
def get_sections():
    """Get unique factory sections."""
    sections = db.session.query(Equipment.factory_section).distinct().all()
    return jsonify([s[0] for s in sections if s[0]])


@api.route('/equipment-names', methods=['GET'])
def get_equipment_names():
    """Get unique equipment names."""
    names = db.session.query(Equipment.equipment_name).distinct().all()
    return jsonify([n[0] for n in names if n[0]])


@api.route('/lines', methods=['GET'])
def get_lines():
    """Get unique lines."""
    lines = db.session.query(Equipment.line).distinct().all()
    return jsonify([l[0] for l in lines if l[0]])


# ==================== HELPER FUNCTIONS ====================

def _get_overdue_equipment_list():
    """Get list of equipment with overdue tasks for dashboard."""
    today = date.today()
    overdue_tasks = PMTask.query.filter(
        PMTask.next_due_date < today,
        PMTask.status != 'Done'
    ).all()
    
    eq_map = {}
    for task in overdue_tasks:
        eq = Equipment.query.get(task.equipment_db_id)
        if eq and eq.id not in eq_map:
            days_overdue = (today - task.next_due_date).days
            eq_map[eq.id] = {
                'id': eq.id,
                'equipment_id': eq.equipment_id,
                'equipment_name': eq.equipment_name,
                'section': eq.factory_section,
                'line': eq.line,
                'overdue_count': eq.overdue_pm_count,
                'max_overdue_days': days_overdue,
                'status': eq.overall_pm_status
            }
        elif eq and eq.id in eq_map:
            days_overdue = (today - task.next_due_date).days
            if days_overdue > eq_map[eq.id]['max_overdue_days']:
                eq_map[eq.id]['max_overdue_days'] = days_overdue
    
    result = list(eq_map.values())
    result.sort(key=lambda x: x['max_overdue_days'], reverse=True)
    return result


def _recalculate_equipment_status(equipment):
    """Recalculate overall PM status and overdue count for an equipment."""
    tasks = PMTask.query.filter_by(equipment_db_id=equipment.id).all()
    today = date.today()
    overdue_count = 0
    max_overdue_days = 0

    for task in tasks:
        if task.next_due_date and not task.actual_done_date:
            days_overdue = (today - task.next_due_date).days
            if days_overdue > 0:
                overdue_count += 1
                max_overdue_days = max(max_overdue_days, days_overdue)

    equipment.overdue_pm_count = overdue_count

    if max_overdue_days > 7:
        equipment.overall_pm_status = 'Red'
    elif max_overdue_days > 3:
        equipment.overall_pm_status = 'Orange'
    else:
        equipment.overall_pm_status = 'Green'


# ==================== BULK TASK CREATION ====================

@api.route('/equipment/<int:equipment_id>/tasks/bulk', methods=['POST'])
def create_tasks_bulk(equipment_id):
    """Create multiple PM tasks at once for an equipment."""
    equipment = Equipment.query.get_or_404(equipment_id)
    data = request.get_json()
    tasks_data = data.get('tasks', [])

    if not tasks_data:
        return jsonify({'error': 'No tasks provided'}), 400

    max_task = db.session.query(db.func.max(PMTask.task_no)).filter_by(
        equipment_db_id=equipment_id
    ).scalar() or 0

    created = []
    for i, t in enumerate(tasks_data):
        task = PMTask(
            equipment_db_id=equipment_id,
            task_no=max_task + i + 1,
            pm_task_description=t.get('pm_task_description', ''),
            frequency_days=t.get('frequency_days'),
            tolerance_days=t.get('tolerance_days', 0),
            next_due_date=datetime.strptime(t['next_due_date'], '%Y-%m-%d').date() if t.get('next_due_date') else None,
            status='Pending',
            consumables=t.get('consumables', ''),
            spare_parts=t.get('spare_parts', ''),
            done_by=t.get('done_by', ''),
            verified_by=t.get('verified_by', ''),
            remarks=t.get('remarks', '')
        )
        db.session.add(task)
        created.append(task)

    _recalculate_equipment_status(equipment)
    db.session.commit()
    return jsonify({'message': f'Created {len(created)} tasks', 'tasks': [t.to_dict() for t in created]}), 201


# ==================== EXCEL IMPORT ====================

@api.route('/import-excel', methods=['POST'])
def import_excel():
    """Import PM tasks from uploaded Excel file. Reads all equipment sheets and imports task data."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Empty filename'}), 400

    # Save to temp file
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsm')
    file.save(tmp.name)
    tmp.close()

    try:
        wb = openpyxl.load_workbook(tmp.name, data_only=True)
        imported_count = 0
        skipped_sheets = []

        for sheet_name in wb.sheetnames:
            if sheet_name in ('Master', 'Dashboard', 'PM_History'):
                continue

            ws = wb[sheet_name]
            if ws.max_row <= 3:
                # Only headers, no data rows
                skipped_sheets.append(sheet_name)
                continue

            # Get equipment_id from B1
            eq_id = ws['B1'].value
            if not eq_id:
                skipped_sheets.append(sheet_name)
                continue

            # Find equipment in DB
            equipment = Equipment.query.filter_by(equipment_id=eq_id).first()
            if not equipment:
                skipped_sheets.append(sheet_name)
                continue

            # Read task rows starting from row 4
            max_task = db.session.query(db.func.max(PMTask.task_no)).filter_by(
                equipment_db_id=equipment.id
            ).scalar() or 0

            for row_num in range(4, ws.max_row + 1):
                task_no = ws.cell(row=row_num, column=1).value
                description = ws.cell(row=row_num, column=2).value

                if not description:
                    continue

                freq = ws.cell(row=row_num, column=3).value
                tol = ws.cell(row=row_num, column=4).value
                next_due = ws.cell(row=row_num, column=5).value
                status = ws.cell(row=row_num, column=6).value
                done_date = ws.cell(row=row_num, column=7).value
                consumables = ws.cell(row=row_num, column=8).value
                spare_parts = ws.cell(row=row_num, column=9).value
                done_by = ws.cell(row=row_num, column=10).value
                verified_by = ws.cell(row=row_num, column=11).value
                remarks = ws.cell(row=row_num, column=12).value

                # Parse dates
                next_due_date = None
                actual_done_date = None
                if isinstance(next_due, datetime):
                    next_due_date = next_due.date()
                elif isinstance(next_due, date):
                    next_due_date = next_due
                if isinstance(done_date, datetime):
                    actual_done_date = done_date.date()
                elif isinstance(done_date, date):
                    actual_done_date = done_date

                max_task += 1
                task = PMTask(
                    equipment_db_id=equipment.id,
                    task_no=task_no if task_no else max_task,
                    pm_task_description=str(description),
                    frequency_days=int(freq) if freq else None,
                    tolerance_days=int(tol) if tol else 0,
                    next_due_date=next_due_date,
                    status=str(status) if status else 'Pending',
                    actual_done_date=actual_done_date,
                    consumables=str(consumables) if consumables else '',
                    spare_parts=str(spare_parts) if spare_parts else '',
                    done_by=str(done_by) if done_by else '',
                    verified_by=str(verified_by) if verified_by else '',
                    remarks=str(remarks) if remarks else ''
                )
                db.session.add(task)
                imported_count += 1

            _recalculate_equipment_status(equipment)

        # Also import PM_History sheet if it has data
        history_count = 0
        if 'PM_History' in wb.sheetnames:
            ws_hist = wb['PM_History']
            for row_num in range(2, ws_hist.max_row + 1):
                eq_id_val = ws_hist.cell(row=row_num, column=1).value
                if not eq_id_val:
                    continue

                old_due = ws_hist.cell(row=row_num, column=6).value
                done_dt = ws_hist.cell(row=row_num, column=7).value
                comp_ts = ws_hist.cell(row=row_num, column=13).value

                old_due_date = None
                actual_done = None
                comp_timestamp = None

                if isinstance(old_due, datetime):
                    old_due_date = old_due.date()
                elif isinstance(old_due, date):
                    old_due_date = old_due

                if isinstance(done_dt, datetime):
                    actual_done = done_dt.date()
                elif isinstance(done_dt, date):
                    actual_done = done_dt

                if isinstance(comp_ts, datetime):
                    comp_timestamp = comp_ts

                history = PMHistory(
                    equipment_id=str(eq_id_val),
                    task_no=ws_hist.cell(row=row_num, column=2).value,
                    pm_task=str(ws_hist.cell(row=row_num, column=3).value or ''),
                    frequency_days=ws_hist.cell(row=row_num, column=4).value,
                    tolerance_days=ws_hist.cell(row=row_num, column=5).value,
                    old_next_due_date=old_due_date,
                    actual_done_date=actual_done,
                    consumables=str(ws_hist.cell(row=row_num, column=8).value or ''),
                    spare_parts=str(ws_hist.cell(row=row_num, column=9).value or ''),
                    done_by=str(ws_hist.cell(row=row_num, column=10).value or ''),
                    verified_by=str(ws_hist.cell(row=row_num, column=11).value or ''),
                    remarks=str(ws_hist.cell(row=row_num, column=12).value or ''),
                    completion_timestamp=comp_timestamp or datetime.utcnow()
                )
                db.session.add(history)
                history_count += 1

        db.session.commit()
        wb.close()

        return jsonify({
            'message': f'Import complete',
            'tasks_imported': imported_count,
            'history_imported': history_count,
            'skipped_sheets': len(skipped_sheets)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        os.unlink(tmp.name)


# ==================== EXPORT EXCEL ====================

@api.route('/export-excel', methods=['GET'])
def export_excel():
    """Export all PM data to downloadable Excel file."""
    from flask import send_file

    wb = openpyxl.Workbook()

    # Master Sheet
    ws_master = wb.active
    ws_master.title = 'Master'
    ws_master.append(['Factory Section', 'Equipment Name', 'Line', 'Equipment ID',
                       'Equipment Criticality', 'Overall PM Status', 'Overdue PM Count'])

    for eq in Equipment.query.order_by(Equipment.id).all():
        ws_master.append([
            eq.factory_section, eq.equipment_name, eq.line, eq.equipment_id,
            eq.equipment_criticality, eq.overall_pm_status, eq.overdue_pm_count
        ])

    # Equipment sheets with tasks
    for eq in Equipment.query.order_by(Equipment.id).all():
        safe_name = eq.equipment_id[:31]  # Excel sheet name max 31 chars
        ws = wb.create_sheet(title=safe_name)
        ws['A1'] = 'Equipment ID'
        ws['B1'] = eq.equipment_id

        ws.append([])  # blank row 2
        ws.append(['Task No', 'PM Task Description', 'Frequency (Days)', 'Tolerance (Days)',
                    'Next Due Date', 'Status', 'Actual Done Date', 'Consumables',
                    'Spare Parts', 'Done By', 'Verified By', 'Remarks / Observation'])

        tasks = PMTask.query.filter_by(equipment_db_id=eq.id).order_by(PMTask.task_no).all()
        for task in tasks:
            ws.append([
                task.task_no, task.pm_task_description, task.frequency_days,
                task.tolerance_days, task.next_due_date, task.status,
                task.actual_done_date, task.consumables, task.spare_parts,
                task.done_by, task.verified_by, task.remarks
            ])

    # PM History Sheet
    ws_hist = wb.create_sheet(title='PM_History')
    ws_hist.append(['Equipment ID', 'Task No', 'PM Task', 'Frequency (Days)',
                     'Tolerance (Days)', 'Old Next Due Date', 'Actual Done Date',
                     'Consumables', 'Spare Parts', 'Done By', 'Verified By',
                     'Remarks', 'Completion Timestamp'])

    for h in PMHistory.query.order_by(PMHistory.completion_timestamp.desc()).all():
        ws_hist.append([
            h.equipment_id, h.task_no, h.pm_task, h.frequency_days,
            h.tolerance_days, h.old_next_due_date, h.actual_done_date,
            h.consumables, h.spare_parts, h.done_by, h.verified_by,
            h.remarks, h.completion_timestamp
        ])

    # Save to temp
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    wb.save(tmp.name)
    tmp.close()

    return send_file(tmp.name, as_attachment=True,
                     download_name='PM_Tracker_Export.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
